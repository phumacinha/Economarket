from openpyxl import load_workbook
from openpyxl import Workbook
import os
import time

def print_titulo_adicionando_produto_ao_mercado(mercado:str):
    print('\033[42m'+'\033[30m'+'Adicionando produto ao mercado '+'\033[1m'+mercado+'\033[0;0m')

def cabecalho (nome:str):
    print()
    print(nome+':')
    print('-'*50)

def print_menu(nome:str, lista:list, entrada:str=None) -> int:
    cabecalho(nome)

    for k, v in enumerate(lista):
        print(k+1, '-', v)

    print('-'*50)

    if entrada is not None:
        codigo = int(input('\n{}: '.format(entrada)))
        while codigo < 0 or codigo > len(lista):
            codigo = int(input('Entrada inválida! Tente novamente: '))
        
        return codigo

dados = load_workbook('dados.xlsx')
mercados = [mercado[0].value for mercado in dados['mercado']]
produtos = [produto[0].value for produto in dados['produto']]
marcas = [marca[0].value for marca in dados['marca'] if marca[0].value is not None]
itens = []

for item in dados['item']:
    if item[0].value is None:
        continue
    itens.append([item[0].value, item[1].value, item[2].value, item[3].value])


print(marcas, itens)

id_mercado = print_menu('Mercados', mercados, 'Selecione um código de mercado (0 para finalizar)')

while id_mercado != 0:
    os.system('cls' if os.name == 'nt' else 'clear')
    print_titulo_adicionando_produto_ao_mercado(mercados[id_mercado-1])
    
    id_produto = print_menu('Produtos', produtos, 'Selecione um código de produto (0 para trocar de mercado)')
    if id_produto == 0:
        os.system('cls' if os.name == 'nt' else 'clear')
        id_mercado = print_menu('Mercados', mercados, 'Selecione um código de mercado (0 para finalizar)')
        continue;

    os.system('cls' if os.name == 'nt' else 'clear')
    print_titulo_adicionando_produto_ao_mercado(mercados[id_mercado-1])
    print('Produto selecionado: ' + produtos[id_produto-1])
    
    print_menu('Marca', marcas)
    while True:
        marca = input("Selecione um código de marca ou digite um nome para adicionar uma nova marca: ")
        
        if marca.isdigit():
            id_marca = int(marca)
            if id_marca < 1 or id_marca > len(marcas):
                print('Código de marca inválido!')
                continue
        else:    
            marcas.append(marca)
            id_marca = len(marcas)
            dados['marca'].cell(row=id_marca, column=1).value = marca 

        break
    
    valor = '%.2f'%float(input('\nValor do produto: '))

    itens.append([id_mercado, id_produto, id_marca, valor])
    id_item = len(itens)
    dados['item'].cell(row=id_item, column=1).value = id_mercado 
    dados['item'].cell(row=id_item, column=2).value = id_produto 
    dados['item'].cell(row=id_item, column=3).value = id_marca 
    dados['item'].cell(row=id_item, column=4).value = valor
    dados.save('dados.xlsx')
    
    os.system('cls' if os.name == 'nt' else 'clear')
    print_titulo_adicionando_produto_ao_mercado(mercados[id_mercado-1])

    print('\033[32m'+'\033[1m'+'Produto adicionado com sucesso!'+'\033[0;0m')
    time.sleep(.6)

